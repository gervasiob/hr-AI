import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin
from urllib.request import Request, urlopen

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from recruitment.models import RemoteTableRecord, RemoteTableSync


class RemoteTableSyncService:
    PAGINATED_API_TABLES = {
        "candidate": "candidates/",
    }
    JSON_API_TABLES = {
        "cvfile": "cv-files/",
    }
    BULK_BATCH_SIZE = 1000

    def list_available_tables(self) -> list[str]:
        payload = self._get_json("tables/")
        tables = payload.get("tables", [])
        return sorted(str(table).strip() for table in tables if str(table).strip())

    def sync_tables(self, table_names: list[str], reset: bool = False) -> list[dict]:
        results = []
        for table_name in table_names:
            result = self.sync_table(table_name, reset=reset)
            results.append(result)
        return results

    def sync_table(self, table_name: str, reset: bool = False) -> dict:
        tracker = self._get_tracker(table_name)

        try:
            if reset:
                self.reset_table(table_name)
                tracker.refresh_from_db()

            rows = self._fetch_table_rows(table_name)
            created_count = 0
            updated_count = 0
            max_remote_id = tracker.last_remote_id
            existing_records = {
                record.remote_id: record
                for record in tracker.records.all().only("id", "remote_id", "payload")
            }
            to_create = []
            to_update = []

            with transaction.atomic():
                current_time = timezone.now()
                for row in rows:
                    remote_id = self._extract_remote_id(row)
                    if remote_id is None:
                        continue

                    if remote_id > max_remote_id:
                        max_remote_id = remote_id

                    existing_record = existing_records.get(remote_id)
                    if remote_id <= tracker.last_remote_id and existing_record and existing_record.payload == row:
                        continue

                    if existing_record is None:
                        to_create.append(
                            RemoteTableRecord(
                                table=tracker,
                                remote_id=remote_id,
                                payload=row,
                                created_at=current_time,
                                updated_at=current_time,
                            )
                        )
                        created_count += 1
                    else:
                        existing_record.payload = row
                        existing_record.updated_at = current_time
                        to_update.append(existing_record)
                        updated_count += 1

                if to_create:
                    RemoteTableRecord.objects.bulk_create(
                        to_create,
                        batch_size=self.BULK_BATCH_SIZE,
                    )

                if to_update:
                    RemoteTableRecord.objects.bulk_update(
                        to_update,
                        ["payload", "updated_at"],
                        batch_size=self.BULK_BATCH_SIZE,
                    )

                tracker.last_remote_id = max_remote_id
                tracker.records_count = len(existing_records) + created_count
                tracker.last_synced_at = timezone.now()
                tracker.last_sync_status = RemoteTableSync.SyncStatus.SUCCESS
                tracker.last_error = ""
                tracker.endpoint = self._resolve_tracker_endpoint(table_name)
                tracker.save(
                    update_fields=[
                        "last_remote_id",
                        "records_count",
                        "last_synced_at",
                        "last_sync_status",
                        "last_error",
                        "endpoint",
                        "updated_at",
                    ]
                )

            return {
                "table_name": table_name,
                "created_count": created_count,
                "updated_count": updated_count,
                "records_count": tracker.records_count,
                "last_remote_id": tracker.last_remote_id,
            }
        except Exception as exc:
            tracker.last_sync_status = RemoteTableSync.SyncStatus.FAILED
            tracker.last_error = str(exc)
            tracker.last_synced_at = timezone.now()
            tracker.save(
                update_fields=[
                    "last_sync_status",
                    "last_error",
                    "last_synced_at",
                    "updated_at",
                ]
            )
            raise

    def reset_table(self, table_name: str) -> None:
        tracker = self._get_tracker(table_name)
        tracker.records.all().delete()
        tracker.last_remote_id = 0
        tracker.records_count = 0
        tracker.last_synced_at = None
        tracker.last_sync_status = RemoteTableSync.SyncStatus.IDLE
        tracker.last_error = ""
        tracker.save(
            update_fields=[
                "last_remote_id",
                "records_count",
                "last_synced_at",
                "last_sync_status",
                "last_error",
                "updated_at",
            ]
        )

    def reset_tables(self, table_names: list[str]) -> None:
        for table_name in table_names:
            self.reset_table(table_name)

    def sync_and_reset_tables(self, table_names: list[str]) -> list[dict]:
        return self.sync_tables(table_names, reset=True)

    def update_selected_tables(self, selected_tables: list[str]) -> None:
        selected_set = set(selected_tables)
        for table_name in self.list_available_tables():
            tracker = self._get_tracker(table_name)
            tracker.is_selected = table_name in selected_set
            tracker.save(update_fields=["is_selected", "updated_at"])

    def get_dashboard_data(self) -> dict:
        available_tables = self.list_available_tables()
        trackers = []
        for table_name in available_tables:
            tracker = self._get_tracker(table_name)
            trackers.append(tracker)
        return {
            "available_tables": available_tables,
            "trackers": trackers,
            "selected_tables": [tracker.table_name for tracker in trackers if tracker.is_selected],
            "total_records": sum(tracker.records_count for tracker in trackers),
        }

    def _fetch_table_rows(self, table_name: str) -> list[dict]:
        if table_name in self.PAGINATED_API_TABLES:
            return self._fetch_paginated_api_rows(table_name)
        if table_name in self.JSON_API_TABLES:
            return self._fetch_json_api_rows(table_name)

        content = self._download_binary(f"export/{table_name}/")
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [self._normalize_header(value) for value in rows[0]]
        results = []
        for values in rows[1:]:
            if not any(value is not None and value != "" for value in values):
                continue
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = self._normalize_value(values[index] if index < len(values) else None)
            results.append(row)
        results.sort(key=lambda item: self._extract_remote_id(item) or 0)
        return results

    def _fetch_json_api_rows(self, table_name: str) -> list[dict]:
        endpoint_path = self.JSON_API_TABLES[table_name]
        payload = self._get_json(endpoint_path)
        if not isinstance(payload, list):
            raise RuntimeError(f"La respuesta JSON de '{table_name}' no es una lista.")

        results = []
        for row in payload:
            if not isinstance(row, dict):
                continue
            results.append({key: self._normalize_value(value) for key, value in row.items()})
        results.sort(key=lambda item: self._extract_remote_id(item) or 0)
        return results

    def _fetch_paginated_api_rows(self, table_name: str) -> list[dict]:
        endpoint_path = self.PAGINATED_API_TABLES[table_name]
        next_url = self._build_endpoint(
            f"{endpoint_path}?page=1&page_size={settings.REMOTE_API_PAGE_SIZE}&ordering=id"
        )
        results = []

        while next_url:
            payload = self._get_json_from_url(next_url)
            page_results = payload.get("results", [])
            for row in page_results:
                results.append({key: self._normalize_value(value) for key, value in row.items()})
            next_url = payload.get("next")

        results.sort(key=lambda item: self._extract_remote_id(item) or 0)
        return results

    def _extract_remote_id(self, row: dict) -> int | None:
        value = row.get("id")
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _normalize_header(self, value) -> str:
        return str(value).strip() if value is not None else ""

    def _normalize_value(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bytes):
            return value.decode("utf-8", errors="replace")
        if isinstance(value, (list, dict, str, int, float, bool)):
            return value
        return str(value)

    def _get_tracker(self, table_name: str) -> RemoteTableSync:
        tracker, _ = RemoteTableSync.objects.get_or_create(
            table_name=table_name,
            defaults={"endpoint": self._resolve_tracker_endpoint(table_name)},
        )
        return tracker

    def _resolve_tracker_endpoint(self, table_name: str) -> str:
        if table_name in self.PAGINATED_API_TABLES:
            return self._build_endpoint(self.PAGINATED_API_TABLES[table_name])
        if table_name in self.JSON_API_TABLES:
            return self._build_endpoint(self.JSON_API_TABLES[table_name])
        return self._build_endpoint(f"export/{table_name}/")

    def _get_json(self, relative_path: str) -> dict:
        endpoint = self._build_endpoint(relative_path)
        return self._get_json_from_url(endpoint)

    def _get_json_from_url(self, endpoint: str) -> dict:
        request = Request(endpoint, headers={"Accept": "application/json"})
        try:
            with urlopen(request, timeout=settings.REMOTE_API_TIMEOUT) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            raise RuntimeError(self._format_http_error(exc, endpoint)) from exc
        except URLError as exc:
            raise RuntimeError(f"No se pudo conectar con {endpoint}: {exc.reason}") from exc

    def _download_binary(self, relative_path: str) -> bytes:
        endpoint = self._build_endpoint(relative_path)
        request = Request(endpoint)
        try:
            with urlopen(request, timeout=settings.REMOTE_API_TIMEOUT) as response:
                return response.read()
        except HTTPError as exc:
            raise RuntimeError(self._format_http_error(exc, endpoint)) from exc
        except URLError as exc:
            raise RuntimeError(f"No se pudo conectar con {endpoint}: {exc.reason}") from exc

    def _build_endpoint(self, relative_path: str) -> str:
        base_url = settings.REMOTE_API_BASE_URL.rstrip("/") + "/"
        return urljoin(base_url, relative_path.lstrip("/"))

    def _format_http_error(self, exc: HTTPError, endpoint: str) -> str:
        response_body = ""
        try:
            response_body = exc.read().decode("utf-8", errors="replace").strip()
        except Exception:
            response_body = ""

        if response_body:
            response_body = " | Respuesta: " + response_body[:400]

        return f"HTTP {exc.code} al consultar {endpoint}{response_body}"
